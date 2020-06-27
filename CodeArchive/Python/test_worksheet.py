# Copyright (c) 2010-2017 openpyxl

# test imports
import pytest

from itertools import islice

# package imports
from openpyxl.workbook import Workbook
from openpyxl.cell import Cell
from openpyxl.utils import coordinate_from_string
from openpyxl.comments import Comment
from openpyxl.utils.exceptions import (
    SheetTitleException,
    InsufficientCoordinatesException,
    NamedRangeException
    )


class DummyWorkbook:

    encoding = "UTF-8"

    def __init__(self):
        self.sheetnames = []


@pytest.fixture
def Worksheet():
    from ..worksheet import Worksheet
    return Worksheet


class TestWorksheet:


    def test_path(self, Worksheet):
        ws = Worksheet(Workbook())
        assert ws.path == "/xl/worksheets/sheetNone.xml"

    def test_new_worksheet(self, Worksheet):
        wb = Workbook()
        ws = Worksheet(wb)
        assert ws.parent == wb


    def test_get_cell(self, Worksheet):
        ws = Worksheet(Workbook())
        cell = ws.cell(row=1, column=1)
        assert cell.coordinate == 'A1'


    def test_invalid_cell(self, Worksheet):
        wb = Workbook()
        ws = Worksheet(wb)
        with pytest.raises(ValueError):
            cell = ws.cell(row=0, column=0)


    def test_worksheet_dimension(self, Worksheet):
        ws = Worksheet(Workbook())
        assert 'A1:A1' == ws.calculate_dimension()
        ws['B12'].value = 'AAA'
        assert 'B12:B12' == ws.calculate_dimension()


    def test_squared_range(self, Worksheet):
        ws = Worksheet(Workbook())
        expected = [
            ('A1', 'B1', 'C1'),
            ('A2', 'B2', 'C2'),
            ('A3', 'B3', 'C3'),
            ('A4', 'B4', 'C4'),
        ]
        rows = ws.get_squared_range(1, 1, 3, 4)
        for row, coord in zip(rows, expected):
            assert tuple(c.coordinate for c in row) == coord


    @pytest.mark.parametrize("row, column, coordinate",
                             [
                                 (1, 0, 'A1'),
                                 (9, 2, 'C9'),
                             ])
    def test_fill_rows(self, Worksheet, row, column, coordinate):
        ws = Worksheet(Workbook())
        ws['A1'] = 'first'
        ws['C9'] = 'last'
        assert ws.calculate_dimension() == 'A1:C9'
        rows = ws.iter_rows()
        first_row = next(islice(rows, row - 1, row))
        assert first_row[column].coordinate == coordinate


    def test_iter_rows(self, Worksheet):
        ws = Worksheet(Workbook())
        expected = [
            ('A1', 'B1', 'C1'),
            ('A2', 'B2', 'C2'),
            ('A3', 'B3', 'C3'),
            ('A4', 'B4', 'C4'),
        ]

        rows = ws.iter_rows(min_row=1, min_col=1, max_row=4, max_col=3)
        for row, coord in zip(rows, expected):
            assert tuple(c.coordinate for c in row) == coord


    def test_iter_rows_offset(self, Worksheet):
        ws = Worksheet(Workbook())
        rows = ws.iter_rows(min_row=1, min_col=1, max_row=4, max_col=3,
                            row_offset=1, column_offset=3)
        expected = [
            ('D2', 'E2', 'F2'),
            ('D3', 'E3', 'F3'),
            ('D4', 'E4', 'F4'),
            ('D5', 'E5', 'F5'),
        ]

        for row, coord in zip(rows, expected):
            assert tuple(c.coordinate for c in row) == coord


    def test_get_named_range(self, Worksheet):
        wb = Workbook()
        ws = wb.active
        wb.create_named_range('test_range', ws, value='C5')
        xlrange = tuple(ws.get_named_range('test_range'))
        cell = xlrange[0]
        assert isinstance(cell, Cell)
        assert cell.row == 5


    def test_get_bad_named_range(self, Worksheet):
        ws = Worksheet(Workbook())
        with pytest.raises(KeyError):
            ws.get_named_range('bad_range')


    def test_get_named_range_wrong_sheet(self, Worksheet):
        wb = Workbook()
        ws1 = wb.create_sheet("Sheet1")
        ws2 = wb.create_sheet("Sheet2")
        wb.create_named_range('wrong_sheet_range', ws1, 'C5')
        with pytest.raises(NamedRangeException):
            ws2.get_named_range('wrong_sheet_range')


    def test_cell_alternate_coordinates(self, Worksheet):
        ws = Worksheet(Workbook())
        cell = ws.cell(row=8, column=4)
        assert 'D8' == cell.coordinate

    def test_cell_insufficient_coordinates(self, Worksheet):
        ws = Worksheet(Workbook())
        with pytest.raises(InsufficientCoordinatesException):
            ws.cell(row=8)

    def test_cell_range_name(self):
        wb = Workbook()
        ws = wb.active
        wb.create_named_range('test_range_single', ws, 'B12')
        c_range_name = ws.get_named_range('test_range_single')
        c_cell = ws['B12']
        assert c_range_name == (c_cell,)


    def test_hyperlink_value(self, Worksheet):
        ws = Worksheet(Workbook())
        ws['A1'].hyperlink = "http://test.com"
        assert "http://test.com" == ws['A1'].value
        ws['A1'].value = "test"
        assert "test" == ws['A1'].value


    def test_append(self, Worksheet):
        ws = Worksheet(Workbook())
        ws.append(['value'])
        assert ws['A1'].value == "value"


    def test_append_list(self, Worksheet):
        ws = Worksheet(Workbook())

        ws.append(['This is A1', 'This is B1'])

        assert 'This is A1' == ws['A1'].value
        assert 'This is B1' == ws['B1'].value

    def test_append_dict_letter(self, Worksheet):
        ws = Worksheet(Workbook())

        ws.append({'A' : 'This is A1', 'C' : 'This is C1'})

        assert 'This is A1' == ws['A1'].value
        assert 'This is C1' == ws['C1'].value

    def test_append_dict_index(self, Worksheet):
        ws = Worksheet(Workbook())

        ws.append({1 : 'This is A1', 3 : 'This is C1'})

        assert 'This is A1' == ws['A1'].value
        assert 'This is C1' == ws['C1'].value

    def test_bad_append(self, Worksheet):
        ws = Worksheet(Workbook())
        with pytest.raises(TypeError):
            ws.append("test")


    def test_append_range(self, Worksheet):
        ws = Worksheet(Workbook())
        ws.append(range(30))
        assert ws['AD1'].value == 29


    def test_append_iterator(self, Worksheet):
        def itty():
            for i in range(30):
                yield i

        ws = Worksheet(Workbook())
        gen = itty()
        ws.append(gen)
        assert ws['AD1'].value == 29


    def test_append_2d_list(self, Worksheet):

        ws = Worksheet(Workbook())

        ws.append(['This is A1', 'This is B1'])
        ws.append(['This is A2', 'This is B2'])

        vals = ws.iter_rows(min_row=1, min_col=1, max_row=2, max_col=2)
        expected = (
            ('This is A1', 'This is B1'),
            ('This is A2', 'This is B2'),
        )
        for e, v in zip(expected, ws.values):
            assert e == tuple(v)


    def test_append_cell(self, Worksheet):
        from openpyxl.cell import Cell

        cell = Cell(None, 'A', 1, 25)

        ws = Worksheet(Workbook())
        ws.append([])

        ws.append([cell])

        assert ws['A2'].value == 25


    def test_rows(self, Worksheet):

        ws = Worksheet(Workbook())

        ws['A1'] = 'first'
        ws['C9'] = 'last'

        rows = tuple(ws.rows)

        assert len(rows) == 9
        first_row = rows[0]
        last_row = rows[-1]

        assert first_row[0].value == 'first' and first_row[0].coordinate == 'A1'
        assert last_row[-1].value == 'last'


    def test_no_rows(self, Worksheet):
        ws = Worksheet(Workbook())
        assert ws.rows == ()


    def test_no_cols(self, Worksheet):
        ws = Worksheet(Workbook())
        assert tuple(ws.columns) == ()


    def test_one_cell(self, Worksheet):
        ws = Worksheet(Workbook())
        c = ws['A1']
        assert tuple(ws.rows) == tuple(ws.columns) == ((c,),)


    def test_by_col(self, Worksheet):
        ws = Worksheet(Workbook())
        c = ws['A1']
        cols = ws._cells_by_col(1, 1, 1, 1)
        assert tuple(cols) == ((c,),)


    def test_cols(self, Worksheet):
        ws = Worksheet(Workbook())

        ws['A1'] = 'first'
        ws['C9'] = 'last'
        expected = [
            ('A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8', 'A9'),
            ('B1', 'B2', 'B3', 'B4', 'B5', 'B6', 'B7', 'B8', 'B9'),
            ('C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'C7', 'C8', 'C9'),

        ]

        cols = tuple(ws.columns)
        for col, coord in zip(cols, expected):
            assert tuple(c.coordinate for c in col) == coord

        assert len(cols) == 3

        assert cols[0][0].value == 'first'
        assert cols[-1][-1].value == 'last'


    def test_values(self, Worksheet):
        ws = Worksheet(Workbook())
        ws.append([1, 2, 3])
        ws.append([4, 5, 6])
        vals = ws.values
        assert next(vals) == (1, 2, 3)
        assert next(vals) == (4, 5, 6)


    def test_auto_filter(self, Worksheet):
        ws = Worksheet(Workbook())

        ws.auto_filter.ref = 'c1:g9'
        assert ws.auto_filter.ref == 'C1:G9'

    def test_getitem(self, Worksheet):
        ws = Worksheet(Workbook())
        c = ws['A1']
        assert isinstance(c, Cell)
        assert c.coordinate == "A1"
        assert ws['A1'].value is None


    @pytest.mark.parametrize("key", [
        slice(None, None),
        slice(None, -1),
        ":",
        ]
    )
    def test_getitem_invalid(self, Worksheet, key):
        ws = Worksheet(Workbook())
        with pytest.raises(IndexError):
            c = ws[key]


    def test_setitem(self, Worksheet):
        ws = Worksheet(Workbook())
        ws['A12'] = 5
        assert ws['A12'].value == 5

    def test_getslice(self, Worksheet):
        ws = Worksheet(Workbook())
        ws['B2'] = "cell"
        cell_range = ws['A1':'B2']
        assert cell_range == (
            (ws['A1'], ws['B1']),
            (ws['A2'], ws['B2'])
        )

    @pytest.mark.parametrize("key", ["C", "C:C"])
    def test_get_single__column(self, Worksheet, key):
        ws = Worksheet(Workbook())
        c1 = ws.cell(row=1, column=3)
        c2 = ws.cell(row=2, column=3, value=5)
        assert ws["C"] == (c1, c2)


    @pytest.mark.parametrize("key", [2, "2", "2:2"])
    def test_get_row(self, Worksheet, key):
        ws = Worksheet(Workbook())
        a2 = ws.cell(row=2, column=1)
        b2 = ws.cell(row=2, column=2)
        c2 = ws.cell(row=2, column=3, value=5)
        assert ws[key] == (a2, b2, c2)


    def test_freeze(self, Worksheet):
        ws = Worksheet(Workbook())
        ws.freeze_panes = ws['b2']
        assert ws.freeze_panes == 'B2'

        ws.freeze_panes = ''
        assert ws.freeze_panes is None

        ws.freeze_panes = 'C5'
        assert ws.freeze_panes == 'C5'

        ws.freeze_panes = ws['A1']
        assert ws.freeze_panes is None


    def test_merged_cells_lookup(self, Worksheet):
        ws = Worksheet(Workbook())
        ws._merged_cells.append("A1:N50")
        merged = ws.merged_cells
        assert 'A1' in merged
        assert 'N50' in merged
        assert 'A51' not in merged
        assert 'O1' not in merged


    def test_merged_cell_ranges(self, Worksheet):
        ws = Worksheet(Workbook())
        assert ws.merged_cell_ranges == []


    def test_merge_range_string(self, Worksheet):
        ws = Worksheet(Workbook())
        ws['A1'] = 1
        ws['D4'] = 16
        assert (4, 4) in ws._cells
        ws.merge_cells(range_string="A1:D4")
        assert ws._merged_cells == ["A1:D4"]
        assert (4, 4) not in ws._cells
        assert (1, 1) in ws._cells


    def test_merge_coordinate(self, Worksheet):
        ws = Worksheet(Workbook())
        ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=4)
        assert ws._merged_cells == ["A1:D4"]


    def test_unmerge_range_string(self, Worksheet):
        ws = Worksheet(Workbook())
        ws._merged_cells = ["A1:D4"]
        ws.unmerge_cells("A1:D4")


    def test_unmerge_coordinate(self, Worksheet):
        ws = Worksheet(Workbook())
        ws._merged_cells = ["A1:D4"]
        ws.unmerge_cells(start_row=1, start_column=1, end_row=4, end_column=4)


    @pytest.mark.parametrize("value, result, rows_cols",
                             [
                                 (3, "1:3", None),
                                 (4, "A:D", "cols")
                             ])
    def test_print_title_old(self, value, result, rows_cols):
        wb = Workbook()
        ws = wb.active
        ws.add_print_title(value, rows_cols)
        assert ws.print_titles == result


    @pytest.mark.parametrize("rows, cols, titles",
                             [
                                ("1:4", None, "1:4"),
                                (None, "A:F", "A:F"),
                                ("1:2", "C:D", "1:2,C:D"),
                             ]
                             )
    def test_print_titles_new(self, rows, cols, titles):
        wb = Workbook()
        ws = wb.active
        ws.print_title_rows = rows
        ws.print_title_cols = cols
        assert ws.print_titles == titles


    @pytest.mark.parametrize("cell_range, result",
                             [
                                 ("A1:F5",  ["$A$1:$F$5"]),
                                 (["$A$1:$F$5"],  ["$A$1:$F$5"]),
                             ]
                             )
    def test_print_area(self, cell_range, result):
        wb = Workbook()
        ws = wb.active
        ws.print_area = cell_range
        assert ws.print_area == result


class TestPositioning(object):
    def test_point(self):
        wb = Workbook()
        ws = wb.active
        assert ws.point_pos(top=40, left=150), ('C' == 3)


    @pytest.mark.parametrize("value", ('A1', 'D52', 'X11'))
    def test_roundtrip(self, value):
        wb = Workbook()
        ws = wb.active
        assert ws.point_pos(*ws.cell(value).anchor) == coordinate_from_string(value)


    def test_point_negative(self):
        wb = Workbook()
        ws = wb.active
        with pytest.raises(ValueError):
            assert ws.point_pos(top=-1, left=-1)


def test_freeze_panes_horiz(Worksheet):
    ws = Worksheet(Workbook())
    ws.freeze_panes = 'A4'

    view = ws.sheet_view
    assert len(view.selection) == 1
    assert dict(view.selection[0]) == {'activeCell': 'A1', 'pane': 'bottomLeft', 'sqref': 'A1'}
    assert dict(view.pane) == {'activePane': 'bottomLeft', 'state': 'frozen',
                               'topLeftCell': 'A4', 'ySplit': '3'}


def test_freeze_panes_vert(Worksheet):
    ws = Worksheet(Workbook())
    ws.freeze_panes = 'D1'

    view = ws.sheet_view
    assert len(view.selection) == 1
    assert dict(view.selection[0]) ==  {'activeCell': 'A1', 'pane': 'topRight', 'sqref': 'A1'}
    assert dict(view.pane) == {'activePane': 'topRight', 'state': 'frozen',
                               'topLeftCell': 'D1', 'xSplit': '3'}


def test_freeze_panes_both(Worksheet):
    ws = Worksheet(Workbook())
    ws.freeze_panes = 'D4'

    view = ws.sheet_view
    assert len(view.selection) == 3
    assert dict(view.selection[0]) == {'pane': 'topRight'}
    assert dict(view.selection[1]) == {'pane': 'bottomLeft',}
    assert dict(view.selection[2]) == {'activeCell': 'A1', 'pane': 'bottomRight', 'sqref': 'A1'}
    assert dict(view.pane) == {'activePane': 'bottomRight', 'state': 'frozen',
                               'topLeftCell': 'D4', 'xSplit': '3', "ySplit":"3"}


def test_min_column(Worksheet):
    ws = Worksheet(DummyWorkbook())
    assert ws.min_column == 1


def test_max_column(Worksheet):
    ws = Worksheet(DummyWorkbook())
    ws['F1'] = 10
    ws['F2'] = 32
    ws['F3'] = '=F1+F2'
    ws['A4'] = '=A1+A2+A3'
    assert ws.max_column == 6


def test_min_row(Worksheet):
    ws = Worksheet(DummyWorkbook())
    assert ws.min_row == 1


def test_max_row(Worksheet):
    ws = Worksheet(DummyWorkbook())
    ws.append([])
    ws.append([5])
    ws.append([])
    ws.append([4])
    assert ws.max_row == 4
